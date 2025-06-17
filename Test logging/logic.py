from tkinter import messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import re

measurements = []
expected_entries = 0
current_entry = 0
structured_data = []

from utils import validate_inputs, is_pass

def clear_all_data(entries):
    for entry in entries:
        entry.delete(0, 'end')
    measurements.clear()
    structured_data.clear()
    global current_entry, expected_entries
    current_entry = 0
    expected_entries = 0

def next_entry_data(SS, TP, MV):
    SS.delete(0, 'end')
    TP.delete(0, 'end')
    MV.delete(0, 'end')

def start_entry_process(NUM):
    global expected_entries, current_entry
    try:
        expected_entries = int(NUM.get())
        if expected_entries <= 0:
            raise ValueError
        messagebox.showinfo("Start", f"Ready to enter {expected_entries} test data points.")
        structured_data.clear()
        current_entry = 0
    except ValueError:
        messagebox.showerror("Invalid Input", "Enter a valid number of test points.")

def add_measurement_data(SS, TP, MV):
    global current_entry
    ss = SS.get()
    tp = TP.get()
    mv = MV.get()

    if not ss or not tp or not mv:
        messagebox.showwarning("Missing Input", "Please fill all fields before adding.")
        return

    structured_data.append({"ss": ss, "tp": tp, "mv": mv})
    measurements.append(mv)
    current_entry += 1

    if current_entry < expected_entries:
        messagebox.showinfo("Added", f"Measurement {mv} added.\n{expected_entries - current_entry} remaining.")
        next_entry_data(SS, TP, MV)
    else:
        messagebox.showinfo("Done", f"All {expected_entries} measurements collected.")
        next_entry_data(SS, TP, MV)

def save_data_to_excel(PN, SN, TESTER):
    pno = PN.get()
    sno = SN.get()
    tester = TESTER.get()
    date_str = datetime.now().strftime("%d-%m-%Y")

    if not structured_data or len(structured_data) != expected_entries:
        messagebox.showerror("Incomplete", "Please add all measurements first.")
        return

    errors = validate_inputs(pno, sno, structured_data[0]['ss'], structured_data[0]['tp'], structured_data[0]['mv'])
    if errors:
        messagebox.showerror("Input Error", "\n".join(errors))
        return

    if not tester:
        messagebox.showerror("Missing Tester", "Please enter tester name.")
        return

    file_path = f"{pno}.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Sl No", "Supply/Component", "Probing point (Test point)", "Resistance(Ohm)"])

    header = [cell.value for cell in ws[1]]
    if sno in header:
        col = header.index(sno) + 1
    else:
        col = len(header) + 1
        ws.cell(row=1, column=col).value = sno

    start_row = 2
    for i, data in enumerate(structured_data):
        row = start_row + i
        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=2).value = data["ss"]
        ws.cell(row=row, column=3).value = data["tp"]
        ws.cell(row=row, column=col).value = data["mv"]

    end_row = start_row + len(structured_data)
    result = "PASS" if is_pass(measurements) else "FAIL"
    ws.cell(row=end_row + 1, column=col - 1).value = "Result:"
    ws.cell(row=end_row + 1, column=col).value = result

    ws.cell(row=end_row + 2, column=col - 1).value = "Tested by:"
    ws.cell(row=end_row + 2, column=col).value = tester.upper()

    ws.cell(row=end_row + 3, column=col - 1).value = "Date:"
    ws.cell(row=end_row + 3, column=col).value = date_str

    wb.save(file_path)
    messagebox.showinfo("Saved", f"Data saved to {file_path}")
    measurements.clear()
    structured_data.clear()

import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import re

measurements = []
expected_entries = 0
current_entry = 0
structured_data = {}

root = tk.Tk()
root.title("Cold Test Logger 🧪")
root.geometry("900x750")
root.configure(bg="#f0f0f0")

style = {'font': ('Arial', 13), 'bg': '#f0f0f0'}
entry_style = {'font': ('Arial', 13)}
entries = {}

heading = tk.Label(root, text="Cold Test Logger 🧪", font=('Arial Rounded MT Bold', 20), bg="#f0f0f0", fg="#1F4E78")
heading.pack(pady=12)

frame1 = tk.LabelFrame(root, text="Basic Information", bg="#e8f4fa", font=('Arial', 13, 'bold'), padx=12, pady=12)
frame1.pack(pady=12)

frame2 = tk.LabelFrame(root, text="Actions", bg="#f0f0f0", font=('Arial', 13, 'bold'), padx=12, pady=12)
frame2.pack(pady=12)

entry_wrapper = tk.Frame(root, bg="#f0f0f0")
frame3 = tk.LabelFrame(entry_wrapper, text="Measurement Entry", bg="#f9f9f9", font=('Arial', 13, 'bold'), padx=12, pady=12)
frame4 = tk.LabelFrame(entry_wrapper, text="Options", bg="#f0f0f0", font=('Arial', 13, 'bold'), padx=12, pady=12)

status_label = tk.Label(root, text="Ready to begin ✨", font=('Arial', 12), bg="#f0f0f0", fg="grey")
status_label.pack(side="bottom", pady=5)

def create_input_row(row, label_text, key, parent, placeholder=""):
    label = tk.Label(parent, text=label_text, **style)
    label.grid(row=row, column=0, sticky="e", padx=12, pady=8)
    entry = tk.Entry(parent, **entry_style, width=35)
    entry.insert(0, placeholder)
    entry.grid(row=row, column=1, padx=12, pady=8)
    entries[key] = entry

create_input_row(0, "Part Number", "PN", frame1, "e.g., 146-PCA000045-L")
create_input_row(1, "Serial Number", "SN", frame1, "e.g., AT-5224-13-0003")
create_input_row(2, "Number of Testing Points", "NUM", frame1, "e.g., 5")

entries['PN'].focus()

def clear_all():
    for entry in entries.values():
        entry.delete(0, tk.END)
    measurements.clear()
    structured_data.clear()
    global current_entry, expected_entries
    current_entry = 0
    expected_entries = 0
    status_label.config(text="All fields cleared 🧼")

def clear_fields():
    for key in ['PN', 'SN', 'NUM']:
        if key in entries:
            entries[key].delete(0, tk.END)
    status_label.config(text="Basic fields cleared 🧽")

def start_entry():
    pno = entries['PN'].get()
    sno = entries['SN'].get()
    num = entries['NUM'].get()
    errors = []

    if not re.match(r"^\d{3}-PCA\d{6}-[A-Z]$", pno):
        errors.append("❌ Invalid Part Number (Format: 146-PCA000045-L)")
    if not re.match(r"^AT-\d{4}-\d{2}-\d{4}$", sno):
        errors.append("❌ Invalid Serial Number (Format: AT-5224-13-0003)")
    try:
        n = int(num)
        if n <= 0:
            errors.append("❌ Number of test points must be greater than 0")
    except ValueError:
        errors.append("❌ Please enter a valid integer for test points")

    if errors:
        messagebox.showerror("Oops! Something's wrong...", "\n".join(errors))
        return

    global expected_entries, current_entry
    expected_entries = n
    current_entry = 0
    structured_data.clear()
    measurements.clear()
    status_label.config(text=f"Ready to enter {expected_entries} measurements ⚙️")
    show_extra_fields()

def show_extra_fields():
    frame3.pack_forget()
    frame4.pack_forget()

    create_input_row(0, "Source Supply", "SS", frame3)
    create_input_row(1, "Test Point", "TP", frame3)
    create_input_row(2, "Expected Impedance", "EXP", frame3)
    create_input_row(3, "Measured Value", "MV", frame3)
    create_input_row(4, "Tested By", "TESTER", frame3)
    entry_wrapper.pack(pady=12) 
    frame3.pack(side="left", pady=10)

    tk.Button(frame4, text=" Add Measurement", command=add_measurement,
              font=('Arial', 13), bg='#28a745', fg='white', width=42).grid(row=0, column=0, columnspan=2, padx=10, pady=8)

    tk.Button(frame4, text="Create New File", command=clear_all,
              font=('Arial', 13), bg='#dc3545', fg='white', width=42).grid(row=1, column=0, columnspan=2, pady=12)
    frame4.pack(side="right",pady=10)

def parse_resistance(value):
    match = re.match(r"([\d.]+)\s*([KkMmRr]?)", value)
    if not match:
        return 0
    number, unit = match.groups()
    number = float(number)
    return number * {'K': 1e3, 'M': 1e6, 'R': 1, '': 1}.get(unit.upper(), 0)

def is_pass(measurements):
    for value in measurements:
        if parse_resistance(value) <= 10:
            return False
    return True

def add_measurement():
    global current_entry
    ss = entries['SS'].get()
    tp = entries['TP'].get()
    mv = entries['MV'].get()
    exp = entries['EXP'].get()

    errors = []
    if not ss:
        errors.append("❌ Source Supply is missing")
    if not tp:
        errors.append("❌ Test Point is missing")
    if not re.match(r"^\d+(\.\d+)?[KkMmRr]?$", mv):
        errors.append("❌ Invalid Measured Value (e.g., 4.49K)")

    if errors:
        messagebox.showerror("Hold up!", "\n".join(errors))
        return

    structured_data.setdefault('entries', []).append({"ss": ss, "tp": tp, "mv": mv, "exp": exp})
    measurements.append(mv)
    current_entry += 1

    if current_entry < expected_entries:
        messagebox.showinfo("Added", f"Measurement {mv} added! {expected_entries - current_entry} left.")
        clear_fields()
        status_label.config(text=f"Added {current_entry}/{expected_entries}")
    else:
        messagebox.showinfo("Done ✅", f"All {expected_entries} measurements recorded. Saving to Excel...")
        save_to_excel()
        status_label.config(text="Measurements saved successfully 💾")

def save_to_excel():
    pno = entries['PN'].get()
    sno = entries['SN'].get()
    tester = entries['TESTER'].get()
    date_str = datetime.now().strftime("%d-%m-%Y")

    if not tester:
        messagebox.showerror("Missing Info", "Please enter the tester's name.")
        return

    file_path = f"{pno}.xlsx"
    new_file = not os.path.exists(file_path)

    if new_file:
        wb = Workbook()
        ws = wb.active
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        heading_cell = ws.cell(row=1, column=1)
        heading_cell.value = f"Cold Test - {pno}"
        heading_cell.font = Font(size=14, bold=True, color="1F4E78")
        heading_cell.alignment = Alignment(horizontal="center")

        headers = ["Sl No", "Supply/Component", "Probing Point", "Expected Impedance"]
        ws.append(headers)
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    required_headers = ["Sl No", "Supply/Component", "Probing Point", "Expected Impedance"]
    for idx, header_text in enumerate(required_headers, start=1):
        if ws.cell(row=2, column=idx).value != header_text:
            ws.cell(row=2, column=idx).value = header_text
            ws.cell(row=2, column=idx).fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            ws.cell(row=2, column=idx).font = Font(bold=True)
            ws.cell(row=2, column=idx).alignment = Alignment(horizontal='center')

    current_header = [cell.value for cell in ws[2]]
    if sno not in current_header:
        col = len(current_header) + 1
        ws.cell(row=2, column=col).value = sno
        ws.cell(row=2, column=col).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        ws.cell(row=2, column=col).font = Font(bold=True)
        sn_col = col
    else:
        sn_col = current_header.index(sno) + 1

    for i, data in enumerate(structured_data['entries']):
        ws.cell(row=i+3, column=1).value = i + 1
        ws.cell(row=i+3, column=2).value = data['ss']
        ws.cell(row=i+3, column=3).value = data['tp']
        ws.cell(row=i+3, column=4).value = data['exp']
        ws.cell(row=i+3, column=sn_col).value = f">{data['mv']}"

    end_row = len(structured_data['entries']) + 3
    result = "PASS" if is_pass(measurements) else "FAIL"

    ws.cell(row=end_row, column=sn_col - 1).value = "Result:"
    result_cell = ws.cell(row=end_row, column=sn_col)
    result_cell.value = result
    result_cell.font = Font(bold=True, color="FF0000" if result == "FAIL" else "008000")

    ws.cell(row=end_row+1, column=sn_col - 1).value = "Tested by:"
    ws.cell(row=end_row+1, column=sn_col).value = tester.upper()

    ws.cell(row=end_row+2, column=sn_col - 1).value = "Date:"
    ws.cell(row=end_row+2, column=sn_col).value = date_str

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)
    messagebox.showinfo("Saved ✅", f"Data saved successfully to:\n{file_path}")
    clear_all()

def generate_report_txt():
    pno = entries['PN'].get()
    if not pno:
        messagebox.showerror("Oops", "Please enter the Part Number first to generate a report.")
        return

    file_path = f"{pno}.xlsx"
    if not os.path.exists(file_path):
        messagebox.showerror("File Not Found", f"No file found for Part Number: {pno}")
        return

    wb = load_workbook(file_path)
    ws = wb.active

    sno = entries['SN'].get()
    result = tester = date = "N/A"

    max_row = ws.max_row
    for i in range(max_row - 3, max_row + 1):
        label = ws.cell(row=i, column=ws.max_column - 1).value
        value = ws.cell(row=i, column=ws.max_column).value
        if label == "Result:":
            result = value
        elif label == "Tested by:":
            tester = value
        elif label == "Date:":
            date = value

    report_lines = [
        "*********** Cold Test Report ***********",
        f"Part Number    : {pno}",
        f"Serial Number  : {sno if sno else 'N/A'}",
        f"Tested By      : {tester}",
        f"Date           : {date}",
        f"Final Result   : {result}",
        "***************************************"
    ]

    report_text = "\n".join(report_lines)
    report_filename = f"{pno}_Report.txt"

    with open(report_filename, "w") as f:
        f.write(report_text)

    messagebox.showinfo("Report Generated ✅", f"Text report saved as:\n{report_filename}")

tk.Button(frame2, text=" Start Entry", command=start_entry,
    font=('Arial', 13), bg='#007acc', fg='white', width=20).grid(row=0, column=0, padx=12, pady=8)

tk.Button(frame2, text="Clear Fields", command=clear_fields,
    font=('Arial', 13), bg='#17a2b8', fg='white', width=20).grid(row=0, column=1, padx=12, pady=8)
tk.Button(frame2, text="Generate Report", command=generate_report_txt,
    font=('Arial', 13), bg='#6f42c1', fg='white', width=20).grid(row=0, column=2, padx=12, pady=8)
root.mainloop()

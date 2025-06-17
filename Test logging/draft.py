import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import re

measurements = []
expected_entries = 0
current_entry = 0
structured_data = []

def validate_inputs(pno, sno, ss, tp, mv):
    errors = []
    if not re.match(r"^\d{3}-PCA\d{6}-[A-Z]$", pno):
        errors.append("\u274c Invalid Part Number format (e.g., 146-PCA000045-L)")
    if not re.match(r"^AT-\d{4}-\d{2}-\d{4}$", sno):
        errors.append("\u274c Invalid Serial Number format (e.g., AT-5224-13-0003)")
    if not ss:
        errors.append("\u274c Source Supply cannot be empty")
    if not tp:
        errors.append("\u274c Test Point cannot be empty")
    if not re.match(r"^\d+(\.\d+)?[KkMmRr]?$", mv):
        errors.append("\u274c Invalid Measurement format (e.g., 4.49K)")
    return errors

def is_pass(measurements):
    for value in measurements:
        numbers = re.findall(r"[\d.]+", value)
        if not numbers or float(numbers[0]) <= 10:
            return False
    return True

def clear_all():
    for entry in [PN, SN, SS, TP, MV, TESTER, NUM]:
        entry.delete(0, tk.END)
    measurements.clear()
    structured_data.clear()
    global current_entry, expected_entries
    current_entry = 0
    expected_entries = 0

def next_entry():
    SS.delete(0, tk.END)
    TP.delete(0, tk.END)
    MV.delete(0, tk.END)

def start_entry():
    global expected_entries, current_entry
    try:
        expected_entries = int(NUM.get())
        if expected_entries <= 0:
            raise ValueError
        messagebox.showinfo("Start", f"Ready to enter {expected_entries} test data points.")
        structured_data.clear()
        current_entry = 0
    except ValueError:
        messagebox.showerror("Invalid Input", "Enter a valid number of test points.")

def add_measurement():
    global current_entry
    ss = SS.get()
    tp = TP.get()
    mv = MV.get()

    if not ss or not tp or not mv:
        messagebox.showwarning("Missing Input", "Please fill all fields before adding.")
        return

    structured_data.append({"ss": ss, "tp": tp, "mv": mv})
    measurements.append(mv)
    current_entry += 1

    if current_entry < expected_entries:
        messagebox.showinfo("Added", f"Measurement {mv} added.\n{expected_entries - current_entry} remaining.")
        next_entry()
    else:
        messagebox.showinfo("Done", f"All {expected_entries} measurements collected.")
        next_entry()

def save_to_excel():
    pno = PN.get()
    sno = SN.get()
    tester = TESTER.get()
    date_str = datetime.now().strftime("%d-%m-%Y")

    if not structured_data or len(structured_data) != expected_entries:
        messagebox.showerror("Incomplete", "Please add all measurements first.")
        return

    errors = validate_inputs(pno, sno, structured_data[0]['ss'], structured_data[0]['tp'], structured_data[0]['mv'])
    if errors:
        messagebox.showerror("Input Error", "\n".join(errors))
        return

    if not tester:
        messagebox.showerror("Missing Tester", "Please enter tester name.")
        return

    file_path = f"{pno}.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Sl No", "Supply/Component", "Probing point (Test point)", "Resistance(Ohm)"])

    header = [cell.value for cell in ws[1]]
    if sno in header:
        col = header.index(sno) + 1
    else:
        col = len(header) + 1
        ws.cell(row=1, column=col).value = sno

    start_row = 2
    for i, data in enumerate(structured_data):
        row = start_row + i
        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=2).value = data["ss"]
        ws.cell(row=row, column=3).value = data["tp"]
        ws.cell(row=row, column=col).value = data["mv"]

    end_row = start_row + len(structured_data)
    result = "PASS" if is_pass(measurements) else "FAIL"
    ws.cell(row=end_row + 1, column=col - 1).value = "Result:"
    ws.cell(row=end_row + 1, column=col).value = result

    ws.cell(row=end_row + 2, column=col - 1).value = "Tested by:"
    ws.cell(row=end_row + 2, column=col).value = tester.upper()

    ws.cell(row=end_row + 3, column=col - 1).value = "Date:"
    ws.cell(row=end_row + 3, column=col).value = date_str

    wb.save(file_path)
    messagebox.showinfo("Saved", f"Data saved to {file_path}")
    measurements.clear()
    structured_data.clear()

root = tk.Tk()
root.title("Industrial Test Logger")
root.geometry("650x500")
root.configure(bg="#f0f0f0")
style = {'font': ('Arial', 11), 'bg': '#f0f0f0'}
entry_style = {'font': ('Arial', 11)}
input_frame = tk.Frame(root, bg="#f0f0f0")
input_frame.pack(pady=10)

labels = [
    "Part Number", "Serial Number", "Number of Testing Points",
    "Source Supply", "Test Point", "Measured Value", "Tested By"
]

entries = []
def create_input_row(row, label_text):
    label = tk.Label(input_frame, text=label_text, **style)
    label.grid(row=row, column=0, sticky="e", padx=10, pady=5)
    entry = tk.Entry(input_frame, **entry_style, width=30)
    entry.grid(row=row, column=1, padx=10, pady=5)
    return entry

PN = create_input_row(0, "Part Number")
SN = create_input_row(1, "Serial Number")
NUM = create_input_row(2, "Number of Testing Points")
SS = create_input_row(3, "Source Supply")
TP = create_input_row(4, "Test Point")
MV = create_input_row(5, "Measured Value")
TESTER = create_input_row(6, "Tested By")

# ===== Buttons Frame =====
button_frame = tk.Frame(root, bg="#f0f0f0")
button_frame.pack(pady=20)

tk.Button(button_frame, text="Start Entry", command=start_entry,
          font=('Arial', 11), bg='#007acc', fg='white', width=18).grid(row=0, column=0, padx=10, pady=5)

tk.Button(button_frame, text="Add Measurement", command=add_measurement,
          font=('Arial', 11), bg='#28a745', fg='white', width=18).grid(row=0, column=1, padx=10, pady=5)

tk.Button(button_frame, text="Finish Serial Entry", command=save_to_excel,
          font=('Arial', 11), bg='#ffc107', fg='black', width=18).grid(row=1, column=0, padx=10, pady=5)

tk.Button(button_frame, text="Clear Fields (New SN)", command=next_entry,
          font=('Arial', 11), bg='#17a2b8', fg='white', width=18).grid(row=1, column=1, padx=10, pady=5)

tk.Button(button_frame, text="Create New File", command=clear_all,
          font=('Arial', 11), bg='#dc3545', fg='white', width=38).grid(row=2, column=0, columnspan=2, pady=10)

root.mainloop()

entry_style = {'font': ('Arial', 11)}
